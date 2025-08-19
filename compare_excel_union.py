import os
import re
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# --- Load configuration (prefer config.py; no .env reliance) ---
try:
	import config  # type: ignore
except Exception:
	config = None

def _to_list(value):
	if value is None:
		return []
	if isinstance(value, (list, tuple)):
		return list(value)
	if isinstance(value, str):
		return [s for s in value.split(",")] if value else []
	return [str(value)]

SAMPLE_FILE = getattr(config, "SAMPLE_FILE", None) if config else None
IMR_FILE = getattr(config, "IMR_FILE", None) if config else None
OUTPUT_FILE = getattr(config, "OUTPUT_FILE", None) if config else None
IGNORED_COLUMNS = _to_list(getattr(config, "IGNORED_COLUMNS", [])) if config else []
IGNORED_COLORS = _to_list(getattr(config, "IGNORED_COLORS", [])) if config else []
CELLS_TO_BE_ADDED = _to_list(getattr(config, "CELLS_TO_BE_ADDED", [])) if config else []
IGNORED_RANGES = _to_list(getattr(config, "IGNORED_RANGES", [])) if config else []

# --- CLI overrides ---
try:
	import argparse
	parser = argparse.ArgumentParser(description="Compare two Excel files across the full union of cells")
	parser.add_argument("--sample", dest="sample", default=SAMPLE_FILE, help="Path to sample workbook (.xlsx)")
	parser.add_argument("--imr", dest="imr", default=IMR_FILE, help="Path to IMR workbook (.xlsx)")
	parser.add_argument("--out", dest="out", default=OUTPUT_FILE, help="Path to output workbook (.xlsx)")
	parser.add_argument("--ignored-colors", dest="ignored_colors", default=",".join(IGNORED_COLORS), help="Comma-separated color codes to ignore (e.g., FF00FF00)")
	parser.add_argument("--add-ranges", dest="add_ranges", default=",".join(CELLS_TO_BE_ADDED), help="Comma-separated cell ranges to force-include (e.g., A1:B5,C10:D12)")
	parser.add_argument("--ignore-ranges", dest="ignore_ranges", default=",".join(IGNORED_RANGES), help="Comma-separated cell ranges to ignore (e.g., E1:F10)")
	args, _ = parser.parse_known_args()
	SAMPLE_FILE = args.sample or SAMPLE_FILE
	IMR_FILE = args.imr or IMR_FILE
	OUTPUT_FILE = args.out or OUTPUT_FILE
	IGNORED_COLORS = _to_list(args.ignored_colors)
	CELLS_TO_BE_ADDED = _to_list(args.add_ranges)
	IGNORED_RANGES = _to_list(args.ignore_ranges)
except Exception:
	pass

# Defaults if not provided
if not OUTPUT_FILE:
	OUTPUT_FILE = "comparison_report.xlsx"

# --- Check files exist ---
if not os.path.exists(SAMPLE_FILE):
	raise FileNotFoundError(f"Sample file not found: {SAMPLE_FILE}")
if not os.path.exists(IMR_FILE):
	raise FileNotFoundError(f"IMR file not found: {IMR_FILE}")

# --- Load workbooks ---
wb_sample = load_workbook(SAMPLE_FILE)
wb_imr = load_workbook(IMR_FILE)

# Data-only workbooks for formulas (evaluated results)
wb_sample_data = load_workbook(SAMPLE_FILE, data_only=True)
wb_imr_data = load_workbook(IMR_FILE, data_only=True)


# --- Helper functions ---
def safe_str(value):
	return str(value) if value is not None else "None"

# Normalize color to a consistent string format
# Handles RGB, indexed, theme colors, and plain strings
def normalize_color(color):
	try:
		if not color:
			return "None"
		if hasattr(color, "rgb") and color.rgb:
			return str(color.rgb)
		if hasattr(color, "indexed") and color.indexed is not None:
			return f"indexed:{color.indexed}"
		if hasattr(color, "theme") and color.theme is not None:
			return f"theme:{color.theme}"
		return str(color)
	except Exception as e:
		return f"Error:{e}"

# Compare font properties between two font objects
# Returns a list of issues if there are mismatches
def compare_fonts(f1, f2):
	issues = []
	if (f1.name != f2.name or f1.bold != f2.bold or f1.italic != f2.italic or
			f1.underline != f2.underline or f1.size != f2.size):
		issues.append({
			"type": "Font Mismatch",
			"sample": f"{safe_str(f1.name)} {safe_str(f1.size)} Bold:{safe_str(f1.bold)} "
			          f"Italic:{safe_str(f1.italic)} Underline:{safe_str(f1.underline)}",
			          "generated": f"{safe_str(f2.name)} {safe_str(f2.size)} Bold:{safe_str(f2.bold)} "
			          f"Italic:{safe_str(f2.italic)} Underline:{safe_str(f2.underline)}"
			          })
	return issues

# Compare alignment properties between two alignment objects
# Returns a list of issues if there are mismatches
def compare_alignment(a1, a2):
	issues = []
	if safe_str(a1.horizontal) != safe_str(a2.horizontal) or safe_str(a1.vertical) != safe_str(a2.vertical):
		issues.append({
			"type": "Alignment Mismatch",
			"sample": f"H:{safe_str(a1.horizontal)} V:{safe_str(a1.vertical)}",
			"generated": f"H:{safe_str(a2.horizontal)} V:{safe_str(a2.vertical)}"
		})
	return issues

# Compare fill properties between two fill objects
# Returns a list of issues if there are mismatches
def compare_fill(f1, f2):
	issues = []
	if normalize_color(f1.start_color) != normalize_color(f2.start_color):
		issues.append({
			"type": "Fill Mismatch",
			"sample": f"ColorCode: {normalize_color(f1.start_color)}",
			"generated": f"ColorCode: {normalize_color(f2.start_color)}"
		})
	return issues

# Compare border properties between two border objects
def compare_border(b1, b2):
	issues = []
	if str(b1) != str(b2):
		issues.append({
			"type": "Border Mismatch",
			"sample": str(b1),
			"generated": str(b2)
		})
	return issues

# Compare two cells: format and value
def compare_cell(cell_sample_format, cell_imr_format, cell_sample_val, cell_imr_val):
	issues = []
	if safe_str(cell_sample_val) != safe_str(cell_imr_val):
		issues.append({
			"type": "Value Mismatch",
			"sample": safe_str(cell_sample_val),
			"generated": safe_str(cell_imr_val)
		})
	issues += compare_fonts(cell_sample_format.font, cell_imr_format.font)
	issues += compare_alignment(cell_sample_format.alignment,
								cell_imr_format.alignment)
	issues += compare_fill(cell_sample_format.fill, cell_imr_format.fill)
	issues += compare_border(cell_sample_format.border, cell_imr_format.border)
	return issues

# Check if a cell has an ignored color
def is_ignored_color(cell):
	"""Check if cell has an ignored color"""
	if not IGNORED_COLORS or not IGNORED_COLORS[0]:
		return False

	cell_color = normalize_color(cell.fill.start_color)

	for ignored_color in IGNORED_COLORS:
		ignored_color_clean = ignored_color.strip()
		if ignored_color_clean:
			if cell_color == ignored_color_clean:
				return True
			elif (ignored_color_clean.startswith('FF') and
				  cell_color.endswith(ignored_color_clean[2:])):
				return True
			elif f"theme:{ignored_color_clean}" in cell_color or f"indexed:{ignored_color_clean}" in cell_color:
				return True

	return False

# Get the merged range for a cell, returns (start_row, start_col, end_row, end_col) or None
def get_merged_range(sheet, row, col):
	"""Get the merged range for a cell, returns (start_row, start_col, end_row, end_col) or None"""
	for merged_range in sheet.merged_cells.ranges:
		if (merged_range.min_row <= row <= merged_range.max_row and
				merged_range.min_col <= col <= merged_range.max_col):
			return (merged_range.min_row, merged_range.min_col,
					merged_range.max_row, merged_range.max_col)
	return None

# Return the top-left cell coordinates for a given cell (resolving merged ranges)
def get_top_left_coords(sheet, row, col):
	merged = get_merged_range(sheet, row, col)
	if merged:
		return merged[0], merged[1]
	return row, col

# Read an effective value by resolving merged cells to their top-left value
def get_effective_value(ws_data, ws_format, row, col):
	tl_r, tl_c = get_top_left_coords(ws_format, row, col)
	return ws_data.cell(row=tl_r, column=tl_c).value

# Check if a cell is part of an ignored color range (including merged cells)
def is_cell_in_ignored_range(sheet, row, col):
	"""Check if a cell is part of an ignored color range (including merged cells)"""
	if is_cell_in_added_ranges(row, col):
		return False

	cell = sheet.cell(row=row, column=col)
	if is_ignored_color(cell):
		return True

	merged_range = get_merged_range(sheet, row, col)
	if merged_range:
		start_row, start_col, end_row, end_col = merged_range
		top_left_cell = sheet.cell(row=start_row, column=start_col)
		if is_ignored_color(top_left_cell):
			return True

	return False

# Wrapper: consider ignored color if present in either sheet
# Still honors added ranges and explicit ignored ranges
def is_cell_ignored(ws_sample, ws_imr, row, col):
	if is_cell_in_added_ranges(row, col):
		return False
	if is_cell_in_ignored_ranges(row, col):
		return True
	if is_cell_in_ignored_range(ws_sample, row, col):
		return True
	if is_cell_in_ignored_range(ws_imr, row, col):
		return True
	return False

# Parse a cell range string like 'A1:B5' and return (start_row, start_col, end_row, end_col)
# Uses openpyxl's range_boundaries to handle different formats
# Returns None if the range is invalid or cannot be parsed
def parse_cell_range(cell_range_str):
	"""Parse a cell range string like 'A1:B5' and return (start_row, start_col, end_row, end_col)"""
	try:
		from openpyxl.utils import range_boundaries
		min_col, min_row, max_col, max_row = range_boundaries(cell_range_str)
		return (min_row, min_col, max_row, max_col)
	except Exception as e:
		print(f"Error parsing cell range '{cell_range_str}': {e}")
		return None

# Check if a cell is in any of the specified cell ranges to be added
def is_cell_in_added_ranges(row, col):
	"""Check if a cell is in any of the specified cell ranges to be added"""
	if not CELLS_TO_BE_ADDED or not CELLS_TO_BE_ADDED[0]:
		return False

	for cell_range_str in CELLS_TO_BE_ADDED:
		cell_range_str = cell_range_str.strip()
		if not cell_range_str:
			continue

		parsed_range = parse_cell_range(cell_range_str)
		if parsed_range:
			start_row, start_col, end_row, end_col = parsed_range
			if (start_row <= row <= end_row and start_col <= col <= end_col):
				return True

	return False

# Check if a cell is in any of the specified cell ranges to be ignored
def is_cell_in_ignored_ranges(row, col):
	"""Check if a cell is in any of the specified cell ranges to be ignored"""
	if not IGNORED_RANGES or not IGNORED_RANGES[0]:
		return False

	for cell_range_str in IGNORED_RANGES:
		cell_range_str = cell_range_str.strip()
		if not cell_range_str:
			continue

		parsed_range = parse_cell_range(cell_range_str)
		if parsed_range:
			start_row, start_col, end_row, end_col = parsed_range
			if (start_row <= row <= end_row and start_col <= col <= end_col):
				return True

	return False

# Debug function to print cell colors for troubleshooting
def debug_cell_colors(sheet, max_rows=10, max_cols=10):
	"""Debug function to print cell colors for troubleshooting"""
	print(f"IGNORED_COLORS: {IGNORED_COLORS}")
	print(f"CELLS_TO_BE_ADDED: {CELLS_TO_BE_ADDED}")
	print(f"IGNORED_RANGES: {IGNORED_RANGES}")
	print("Cell colors in first few rows/columns:")
	for r in range(1, min(max_rows + 1, sheet.max_row + 1)):
		for c in range(1, min(max_cols + 1, sheet.max_column + 1)):
			cell = sheet.cell(row=r, column=c)
			cell_color = normalize_color(cell.fill.start_color)
			if cell_color != "None":
				is_ignored = is_ignored_color(cell)
				is_added = is_cell_in_added_ranges(r, c)
				is_ignored_range = is_cell_in_ignored_ranges(r, c)
				print(
					f"Cell {cell.coordinate}: color={cell_color}, ignored={is_ignored}, added={is_added}, ignored_range={is_ignored_range}")

# Find all tables in a sheet based on bordered cells
# Returns a list of tuples (start_row, start_col, end_row, end_col)
def find_bordered_tables(sheet):
	visited = set()
	tables = []
	max_row = sheet.max_row
	max_col = sheet.max_column

	def has_border(cell):
		b = cell.border
		return any([b.left.style, b.right.style, b.top.style, b.bottom.style])

	for i in range(1, max_row + 1):
		for j in range(1, max_col + 1):
			if (i, j) in visited:
				continue
			cell = sheet.cell(row=i, column=j)
			if has_border(cell):
				end_row = i
				while end_row + 1 <= max_row and any(has_border(sheet.cell(row=end_row+1, column=k))
												for k in range(j, max_col+1)):
					end_row += 1
				end_col = j
				while end_col + 1 <= max_col and any(has_border(sheet.cell(row=k, column=end_col+1))
												for k in range(i, end_row+1)):
					end_col += 1
				for r in range(i, end_row + 1):
					for c in range(j, end_col + 1):
						visited.add((r, c))
				tables.append((i, j, end_row, end_col))
	return tables

# Merge rectangles that overlap or touch into union rectangles
def rectangles_overlap_or_touch(a, b):
	a_r1, a_c1, a_r2, a_c2 = a
	b_r1, b_c1, b_r2, b_c2 = b
	# If they are separated by at least one full row/col gap, they do not touch
	if a_r2 < b_r1 - 1:
		return False
	if b_r2 < a_r1 - 1:
		return False
	if a_c2 < b_c1 - 1:
		return False
	if b_c2 < a_c1 - 1:
		return False
	return True


def merge_two_rects(a, b):
	a_r1, a_c1, a_r2, a_c2 = a
	b_r1, b_c1, b_r2, b_c2 = b
	return (
		min(a_r1, b_r1),
		min(a_c1, b_c1),
		max(a_r2, b_r2),
		max(a_c2, b_c2),
	)


def merge_rectangles(rects):
	rects = rects[:]
	changed = True
	while changed and len(rects) > 1:
		changed = False
		new_rects = []
		used = [False] * len(rects)
		for i in range(len(rects)):
			if used[i]:
				continue
			current = rects[i]
			for j in range(i + 1, len(rects)):
				if used[j]:
					continue
				other = rects[j]
				if rectangles_overlap_or_touch(current, other):
					current = merge_two_rects(current, other)
					used[j] = True
					changed = True
			new_rects.append(current)
			used[i] = True
		rects = new_rects
	return rects

import difflib

def header_similarity(text_a, text_b):
	"""Compute similarity between two header strings in [0,1]."""
	if text_a is None or text_b is None:
		return 0.0
	a = str(text_a).strip().lower()
	b = str(text_b).strip().lower()
	if not a or not b:
		return 0.0
	if a == b:
		return 1.0
	return difflib.SequenceMatcher(None, a, b).ratio()

def pair_columns_order_preserving(header_sample_map, header_imr_map, similarity_threshold=0.72):
	"""
	Pair headers from sample and IMR in order-preserving way using fuzzy similarity.
	Inputs are dicts of {header_text: column_index} built in left-to-right order.
	Returns:
	  - paired: list of tuples (sample_header, imr_header)
	  - missing_in_imr: list of sample headers with no match
	  - missing_in_sample: list of imr headers with no match
	"""
	sample_headers = list(header_sample_map.keys())
	imr_headers = list(header_imr_map.keys())

	paired = []
	used_imr = set()

	# First pass: exact matches to lock-in quickly
	for s_hdr in sample_headers:
		if s_hdr in header_imr_map and s_hdr not in used_imr:
			paired.append((s_hdr, s_hdr))
			used_imr.add(s_hdr)

	# Second pass: fuzzy matching for remaining
	for s_hdr in sample_headers:
		if any(s_hdr == p[0] for p in paired):
			continue
		best_hdr = None
		best_score = similarity_threshold
		for i_hdr in imr_headers:
			if i_hdr in used_imr:
				continue
			score = header_similarity(s_hdr, i_hdr)
			if score > best_score:
				best_score = score
				best_hdr = i_hdr
		if best_hdr is not None:
			paired.append((s_hdr, best_hdr))
			used_imr.add(best_hdr)

	# Preserve original left-to-right order based on sample
	paired.sort(key=lambda x: sample_headers.index(x[0]))

	matched_sample = {s for s, _ in paired}
	matched_imr = {i for _, i in paired}
	missing_in_imr = [s for s in sample_headers if s not in matched_sample]
	missing_in_sample = [i for i in imr_headers if i not in matched_imr]

	return paired, missing_in_imr, missing_in_sample

def add_hyperlink(sheet_summary, row, col, target_sheet, display_text=None):
	cell = sheet_summary.cell(row=row, column=col)
	display_text = display_text or target_sheet
	if re.search(r"[^A-Za-z_]", target_sheet):
		safe_sheet_name = f"'{target_sheet}'"
	else:
		safe_sheet_name = target_sheet
	cell.hyperlink = f"#{safe_sheet_name}!A1"
	cell.value = display_text
	# Apply hyperlink formatting: blue color, underline, Aptos Narrow font
	cell.font = Font(name="Aptos Narrow", size=11,
					 color="0000FF", underline="single")
	# Apply border formatting without overriding font
	cell.border = Border(
		left=Side(style='thin'),
		right=Side(style='thin'),
		top=Side(style='thin'),
		bottom=Side(style='thin')
	)
	cell.alignment = Alignment(horizontal="left", vertical="top")

# Apply header formatting to a cell
# Bold, Aptos Narrow font, 12pt size, Light Blue fill, Black text
# Center aligned, with thin borders
def apply_header_formatting(cell):
	"""Apply header formatting: Bold, Aptos Narrow, 12pt, Light Blue fill, Black text, Center aligned"""
	cell.font = Font(name="Aptos Narrow", size=12, bold=True, color="000000")
	cell.fill = PatternFill(start_color="ADD8E6",
							 end_color="ADD8E6", fill_type="solid")
	cell.alignment = Alignment(horizontal="center", vertical="center")
	cell.border = Border(
		left=Side(style='thin'),
		right=Side(style='thin'),
		top=Side(style='thin'),
		bottom=Side(style='thin')
	)

# Apply data formatting to a cell
# Aptos Narrow font, 11pt size, Black text, Top-Left aligned,
# with thin borders
def apply_data_formatting(cell):
	"""Apply data formatting: Aptos Narrow, 11pt, Black text, Top-Left aligned, All borders"""
	cell.font = Font(name="Aptos Narrow", size=11, color="000000")
	cell.alignment = Alignment(horizontal="left", vertical="top")
	cell.border = Border(
		left=Side(style='thin'),
		right=Side(style='thin'),
		top=Side(style='thin'),
		bottom=Side(style='thin')
	)


# --- Main comparison ---
wb_output = Workbook()
summary_ws = wb_output.active
summary_ws.title = "Summary"
summary_ws.append(["Sheet Name", "Hyperlink",
				   "Missing Details", "Issue Count", "Ignored Issues Count"])

# Apply header formatting to summary sheet
for col in range(1, 6):  # 5 columns
	apply_header_formatting(summary_ws.cell(row=1, column=col))
row_summary = 2

# Preserve order: sample sheets first, then IMR-only sheets
ordered_sheetnames = list(wb_sample.sheetnames) + \
	[s for s in wb_imr.sheetnames if s not in wb_sample.sheetnames]

for sheet_name in ordered_sheetnames:
	issues_count = 0
	ignored_issues_count = 0

	# Handle missing sheets
	if sheet_name not in wb_sample.sheetnames:
		summary_ws.cell(row=row_summary, column=1, value=sheet_name)
		add_hyperlink(summary_ws, row_summary, 2, sheet_name)
		summary_ws.cell(row=row_summary, column=3, value="Missing In Sample")
		summary_ws.cell(row=row_summary, column=4, value=0)
		summary_ws.cell(row=row_summary, column=5, value=0)

		apply_data_formatting(summary_ws.cell(row=row_summary, column=1))
		apply_data_formatting(summary_ws.cell(row=row_summary, column=3))
		apply_data_formatting(summary_ws.cell(row=row_summary, column=4))
		apply_data_formatting(summary_ws.cell(row=row_summary, column=5))
		row_summary += 1
		continue

	if sheet_name not in wb_imr.sheetnames:
		summary_ws.cell(row=row_summary, column=1, value=sheet_name)
		add_hyperlink(summary_ws, row_summary, 2, sheet_name)
		summary_ws.cell(row=row_summary, column=3, value="Missing In IMR_Report")
		summary_ws.cell(row=row_summary, column=4, value=0)
		summary_ws.cell(row=row_summary, column=5, value=0)

		apply_data_formatting(summary_ws.cell(row=row_summary, column=1))
		apply_data_formatting(summary_ws.cell(row=row_summary, column=3))
		apply_data_formatting(summary_ws.cell(row=row_summary, column=4))
		apply_data_formatting(summary_ws.cell(row=row_summary, column=5))
		row_summary += 1
		continue

	ws_sample = wb_sample[sheet_name]
	ws_imr = wb_imr[sheet_name]
	ws_sample_data = wb_sample_data[sheet_name]
	ws_imr_data = wb_imr_data[sheet_name]

	ws_out = wb_output.create_sheet(title=sheet_name)
	ws_out.append(["Cell", "Name", "Column Name", "Issue Type",
				   "Sample Value", "Generated Value"])

	for col in range(1, 7):  # 6 columns
		apply_header_formatting(ws_out.cell(row=1, column=col))

	ignored_issues = []
	seen_pairs = set()

	def is_top_left_position(r, c):
		s_r, s_c = get_top_left_coords(ws_sample, r, c)
		i_r, i_c = get_top_left_coords(ws_imr, r, c)
		return (r == s_r and c == s_c) or (r == i_r and c == i_c)

	def canonical_pair_for(r, c):
		s_r, s_c = get_top_left_coords(ws_sample, r, c)
		i_r, i_c = get_top_left_coords(ws_imr, r, c)
		return (s_r, s_c, i_r, i_c)

	# --- Table comparison (UNION from both sheets) ---
	tables_sample = find_bordered_tables(ws_sample)
	tables_imr = find_bordered_tables(ws_imr)
	union_tables = merge_rectangles(tables_sample + tables_imr)

	union_table_cells = set()
	for (start_row, start_col, end_row, end_col) in union_tables:
		for r in range(start_row, end_row + 1):
			for c in range(start_col, end_col + 1):
				union_table_cells.add((r, c))

	for (start_row, start_col, end_row, end_col) in union_tables:
		# Build header maps from header row (exclude the first column which is usually row labels)
		header_sample_map = {}
		header_imr_map = {}
		for c in range(start_col + 1, end_col + 1):
			hs = get_effective_value(ws_sample_data, ws_sample, start_row, c)
			hi = get_effective_value(ws_imr_data, ws_imr, start_row, c)
			if isinstance(hs, str):
				hs = hs.strip()
			if isinstance(hi, str):
				hi = hi.strip()
			if hs not in (None, ""):
				header_sample_map[str(hs)] = c
			if hi not in (None, ""):
				header_imr_map[str(hi)] = c

		# Pair columns using fuzzy, order-preserving alignment
		paired_cols, missing_in_imr, missing_in_sample = pair_columns_order_preserving(header_sample_map, header_imr_map)

		# Report header mismatches as value mismatches (once per pair)
		for s_hdr, i_hdr in paired_cols:
			if s_hdr != i_hdr:
				c_s = header_sample_map[s_hdr]
				issue_data = [
					ws_sample.cell(row=start_row, column=c_s).coordinate,
					"Header",
					s_hdr,
					"Value Mismatch",
					s_hdr,
					i_hdr
				]
				ws_out.append(issue_data)
				issues_count += 1

		# Report true missing columns
		for s_hdr in missing_in_imr:
			c_s = header_sample_map[s_hdr]
			issue_data = [
				ws_sample.cell(row=start_row, column=c_s).coordinate,
				"",
				s_hdr,
				"Missing in report",
				s_hdr,
				""
			]
			ws_out.append(issue_data)
			issues_count += 1

		for i_hdr in missing_in_sample:
			c_i = header_imr_map[i_hdr]
			issue_data = [
				ws_imr.cell(row=start_row, column=c_i).coordinate,
				"",
				i_hdr,
				"Missing in sample",
				"",
				i_hdr
			]
			ws_out.append(issue_data)
			issues_count += 1

		# Now compare data for paired columns row-by-row using the mapped indices
		seen_row_pairs = set()
		for r in range(start_row + 1, end_row + 1):
			# Row header: prefer sample value else IMR
			row_header = get_effective_value(ws_sample_data, ws_sample, r, start_col)
			if row_header is None:
				row_header = get_effective_value(ws_imr_data, ws_imr, r, start_col)

			for s_hdr, i_hdr in paired_cols:
				c_s = header_sample_map[s_hdr]
				c_i = header_imr_map[i_hdr]

				if is_cell_in_ignored_ranges(r, c_s) and is_cell_in_ignored_ranges(r, c_i):
					continue

				# Only evaluate once per (row, sample_col, imr_col)
				pair_key = (r, c_s, c_i)
				if pair_key in seen_row_pairs:
					continue
				seen_row_pairs.add(pair_key)

				# Effective values and formats (respect merged)
				val_sample = get_effective_value(ws_sample_data, ws_sample, r, c_s)
				val_imr = get_effective_value(ws_imr_data, ws_imr, r, c_i)

				s_r, s_c = get_top_left_coords(ws_sample, r, c_s)
				i_r, i_c = get_top_left_coords(ws_imr, r, c_i)
				cell_sample_fmt = ws_sample.cell(row=s_r, column=s_c)
				cell_imr_fmt = ws_imr.cell(row=i_r, column=i_c)

				# Consider ignored if either side is ignored
				is_ignored = (
					is_cell_ignored(ws_sample, ws_imr, r, c_s) or
					is_cell_ignored(ws_sample, ws_imr, r, c_i)
				)

				issues = compare_cell(cell_sample_fmt, cell_imr_fmt, val_sample, val_imr)
				for issue in issues:
					issue_data = [
						ws_sample.cell(row=r, column=c_s).coordinate,
						row_header,
						s_hdr,
						issue["type"],
						issue["sample"],
						issue["generated"]
					]

					if is_ignored:
						ignored_issues.append(issue_data)
						ignored_issues_count += 1
					else:
						current_row = ws_out.max_row + 1
						ws_out.append(issue_data)
						for col in range(1, 7):
							apply_data_formatting(ws_out.cell(row=current_row, column=col))
						issues_count += 1

	# --- Non-table key–value comparison (symmetric across both sheets) ---
	processed_cells = set()
	max_row_union = max(ws_sample.max_row, ws_imr.max_row)
	max_col_union = max(ws_sample.max_column, ws_imr.max_column)

	# Fallback: also sweep the full grid for raw cell-by-cell differences outside tables
	# to catch values in far columns like AAA that are not captured as key-value pairs
	for r in range(1, max_row_union + 1):
		for c in range(1, max_col_union + 1):
			if (r, c) in union_table_cells:
				continue
			if is_cell_in_ignored_ranges(r, c):
				continue
			# Skip cells that will be handled by key-value scanning; only take isolated values
			left_val_s = get_effective_value(ws_sample_data, ws_sample, r, c - 1) if c > 1 else None
			left_val_i = get_effective_value(ws_imr_data, ws_imr, r, c - 1) if c > 1 else None
			this_val_s = get_effective_value(ws_sample_data, ws_sample, r, c)
			this_val_i = get_effective_value(ws_imr_data, ws_imr, r, c)

			# Only consider if at least one side has a value
			if this_val_s is None and this_val_i is None:
				continue

			# De-dup merged regions
			pair_key = canonical_pair_for(r, c)
			if pair_key in seen_pairs:
				continue
			if not is_top_left_position(r, c):
				continue
			seen_pairs.add(pair_key)

			s_r, s_c = get_top_left_coords(ws_sample, r, c)
			i_r, i_c = get_top_left_coords(ws_imr, r, c)
			cell_sample_fmt = ws_sample.cell(row=s_r, column=s_c)
			cell_imr_fmt = ws_imr.cell(row=i_r, column=i_c)

			is_ignored = is_cell_ignored(ws_sample, ws_imr, r, c)

			issues = compare_cell(cell_sample_fmt, cell_imr_fmt, this_val_s, this_val_i)
			for issue in issues:
				issue_data = [
					ws_sample.cell(row=r, column=c).coordinate,
					"",
					"",
					issue["type"],
					issue["sample"],
					issue["generated"]
				]
				if is_ignored:
					ignored_issues.append(issue_data)
					ignored_issues_count += 1
				else:
					current_row = ws_out.max_row + 1
					ws_out.append(issue_data)
					for col in range(1, 7):
						apply_data_formatting(ws_out.cell(row=current_row, column=col))
					issues_count += 1

	def looks_like_key(text):
		if not text or not isinstance(text, str):
			return False
		stripped = text.strip()
		if not stripped:
			return False
		if stripped.isdigit():
			return False
		if stripped.startswith('('):
			return False
		return len(stripped) > 1

	# Scan twice: once using sample keys, once using IMR keys
	for source in ("sample", "imr"):
		src_ws = ws_sample if source == "sample" else ws_imr
		src_data = ws_sample_data if source == "sample" else ws_imr_data
		other_ws = ws_imr if source == "sample" else ws_sample
		other_data = ws_imr_data if source == "sample" else ws_sample_data

		for r in range(1, max_row_union + 1):
			for c in range(1, max_col_union + 1):
				if (r, c) in union_table_cells or (r, c) in processed_cells:
					continue

				key_text = get_effective_value(src_data, src_ws, r, c)
				if not looks_like_key(key_text):
					continue

				# Look for the corresponding value in the next columns on the same row
				for cc in range(c + 1, max_col_union + 1):
					if (r, cc) in union_table_cells:
						break

					if is_cell_in_ignored_ranges(r, cc):
						continue

					# Skip if cc is also a key-like cell to avoid chaining keys
					if looks_like_key(get_effective_value(src_data, src_ws, r, cc)):
						break

					value_src = get_effective_value(src_data, src_ws, r, cc)
					value_other = get_effective_value(other_data, other_ws, r, cc)

					# De-dup against merged canonical pair
					pair_key = canonical_pair_for(r, cc)
					if pair_key in seen_pairs:
						continue
					if not is_top_left_position(r, cc):
						continue
					seen_pairs.add(pair_key)

					# Only consider positions where there is some content (in either)
					if value_src is None and value_other is None:
						continue

					# Effective formats
					s_r, s_c = get_top_left_coords(ws_sample, r, cc)
					i_r, i_c = get_top_left_coords(ws_imr, r, cc)
					cell_sample_fmt = ws_sample.cell(row=s_r, column=s_c)
					cell_imr_fmt = ws_imr.cell(row=i_r, column=i_c)

					is_ignored = is_cell_ignored(ws_sample, ws_imr, r, cc)

					# Map values into sample/imr order expected by compare_cell
					if source == "sample":
						val_sample = value_src
						val_imr = value_other
					else:
						val_sample = value_other
						val_imr = value_src

					issues = compare_cell(cell_sample_fmt, cell_imr_fmt, val_sample, val_imr)
					for issue in issues:
						issue_data = [
							src_ws.cell(row=r, column=cc).coordinate,
							key_text,
							safe_str(value_src) if source == "sample" else safe_str(value_other),
							issue["type"],
							issue["sample"],
							issue["generated"]
						]

						if is_ignored:
							ignored_issues.append(issue_data)
							ignored_issues_count += 1
						else:
							current_row = ws_out.max_row + 1
							ws_out.append(issue_data)
							for col in range(1, 7):
								apply_data_formatting(ws_out.cell(row=current_row, column=col))
							issues_count += 1

					processed_cells.add((r, cc))
					break

				processed_cells.add((r, c))

	# --- Add Ignored Columns Section ---
	if ignored_issues:
		ws_out.append([])
		ws_out.append(["***Columns To Be Ignored***"])
		apply_header_formatting(ws_out.cell(row=ws_out.max_row, column=1))
		ws_out.append(["Cell", "Name", "Column Name", "Issue Type",
					   "Sample Value", "Generated Value"])
		for col in range(1, 7):
			apply_header_formatting(ws_out.cell(row=ws_out.max_row, column=col))
		for issue_data in ignored_issues:
			current_row = ws_out.max_row + 1
			ws_out.append(issue_data)
			for col in range(1, 7):
				apply_data_formatting(ws_out.cell(row=current_row, column=col))

	# --- Summary ---
	summary_ws.cell(row=row_summary, column=1, value=sheet_name)
	add_hyperlink(summary_ws, row_summary, 2, sheet_name)
	summary_ws.cell(row=row_summary, column=3, value="")
	summary_ws.cell(row=row_summary, column=4, value=issues_count)
	summary_ws.cell(row=row_summary, column=5, value=ignored_issues_count)

	apply_data_formatting(summary_ws.cell(row=row_summary, column=1))
	apply_data_formatting(summary_ws.cell(row=row_summary, column=3))
	apply_data_formatting(summary_ws.cell(row=row_summary, column=4))
	apply_data_formatting(summary_ws.cell(row=row_summary, column=5))
	row_summary += 1

# Remove default sheet if empty
if "Sheet" in wb_output.sheetnames and wb_output["Sheet"].max_row == 1:
	del wb_output["Sheet"]

# --- Save output ---
wb_output.save(OUTPUT_FILE)
print(f"Comparison report saved to: {OUTPUT_FILE}")