import os
import re
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from dotenv import load_dotenv

# --- Load environment variables ---
load_dotenv()

SAMPLE_FILE = os.getenv("SAMPLE_FILE")
IMR_FILE = os.getenv("IMR_FILE")
OUTPUT_FILE = os.getenv("OUTPUT_FILE")
IGNORED_COLUMNS = os.getenv("IGNORED_COLUMNS", "").split(",")
IGNORED_COLORS = os.getenv("IGNORED_COLORS", "").split(",")
CELLS_TO_BE_ADDED = os.getenv("CELLS_TO_BE_ADDED", "").split(",")
IGNORED_RANGES = os.getenv("IGNORED_RANGES", "").split(",")

# --- Check files exist ---
if not os.path.exists(SAMPLE_FILE):
    raise FileNotFoundError(f"Sample file not found: {SAMPLE_FILE}")
if not os.path.exists(IMR_FILE):
    raise FileNotFoundError(f"IMR file not found: {IMR_FILE}")

# --- Load workbooks ---
wb_sample = load_workbook(SAMPLE_FILE)
wb_imr = load_workbook(IMR_FILE)

# Data-only workbooks for formulas (evaluated results)
wb_sample_data = load_workbook(SAMPLE_FILE, data_only=True)
wb_imr_data = load_workbook(IMR_FILE, data_only=True)


# --- Helper functions ---
def safe_str(value):
    return str(value) if value is not None else "None"

# Normalize color to a consistent string format
# Handles RGB, indexed, theme colors, and plain strings
def normalize_color(color):
    try:
        if not color:
            return "None"
        if hasattr(color, "rgb") and color.rgb:
            return str(color.rgb)
        if hasattr(color, "indexed") and color.indexed is not None:
            return f"indexed:{color.indexed}"
        if hasattr(color, "theme") and color.theme is not None:
            return f"theme:{color.theme}"
        return str(color)
    except Exception as e:
        return f"Error:{e}"

# Compare font properties between two font objects
# Returns a list of issues if there are mismatches
def compare_fonts(f1, f2):
    issues = []
    if (f1.name != f2.name or f1.bold != f2.bold or f1.italic != f2.italic or
            f1.underline != f2.underline or f1.size != f2.size):
        issues.append({
            "type": "Font Mismatch",
            "sample": f"{safe_str(f1.name)} {safe_str(f1.size)} Bold:{safe_str(f1.bold)} "
                      f"Italic:{safe_str(f1.italic)} Underline:{safe_str(f1.underline)}",
                      "generated": f"{safe_str(f2.name)} {safe_str(f2.size)} Bold:{safe_str(f2.bold)} "
                      f"Italic:{safe_str(f2.italic)} Underline:{safe_str(f2.underline)}"
                      })
    return issues

# Compare alignment properties between two alignment objects
# Returns a list of issues if there are mismatches
def compare_alignment(a1, a2):
    issues = []
    if safe_str(a1.horizontal) != safe_str(a2.horizontal) or safe_str(a1.vertical) != safe_str(a2.vertical):
        issues.append({
            "type": "Alignment Mismatch",
            "sample": f"H:{safe_str(a1.horizontal)} V:{safe_str(a1.vertical)}",
            "generated": f"H:{safe_str(a2.horizontal)} V:{safe_str(a2.vertical)}"
        })
    return issues

# Compare fill properties between two fill objects
# Returns a list of issues if there are mismatches
def compare_fill(f1, f2):
    issues = []
    if normalize_color(f1.start_color) != normalize_color(f2.start_color):
        issues.append({
            "type": "Fill Mismatch",
            "sample": f"ColorCode: {normalize_color(f1.start_color)}",
            "generated": f"ColorCode: {normalize_color(f2.start_color)}"
        })
    return issues

# Compare border properties between two border objects
def compare_border(b1, b2):
    issues = []
    if str(b1) != str(b2):
        issues.append({
            "type": "Border Mismatch",
            "sample": str(b1),
            "generated": str(b2)
        })
    return issues

# Compare two cells: format and value
def compare_cell(cell_sample_format, cell_imr_format, cell_sample_val, cell_imr_val):
    issues = []
    if safe_str(cell_sample_val) != safe_str(cell_imr_val):
        issues.append({
            "type": "Value Mismatch",
            "sample": safe_str(cell_sample_val),
            "generated": safe_str(cell_imr_val)
        })
    issues += compare_fonts(cell_sample_format.font, cell_imr_format.font)
    issues += compare_alignment(cell_sample_format.alignment,
                                cell_imr_format.alignment)
    issues += compare_fill(cell_sample_format.fill, cell_imr_format.fill)
    issues += compare_border(cell_sample_format.border, cell_imr_format.border)
    return issues

# Check if a cell has an ignored color
def is_ignored_color(cell):
    """Check if cell has an ignored color"""
    if not IGNORED_COLORS or not IGNORED_COLORS[0]:
        return False

    cell_color = normalize_color(cell.fill.start_color)

    for ignored_color in IGNORED_COLORS:
        ignored_color_clean = ignored_color.strip()
        if ignored_color_clean:
            if cell_color == ignored_color_clean:
                return True
            elif cell_color.endswith(ignored_color_clean[2:]) and ignored_color_clean.startswith('FF'):
                return True
            elif f"theme:{ignored_color_clean}" in cell_color or f"indexed:{ignored_color_clean}" in cell_color:
                return True

    return False

# Get the merged range for a cell, returns (start_row, start_col, end_row, end_col) or None
def get_merged_range(sheet, row, col):
    """Get the merged range for a cell, returns (start_row, start_col, end_row, end_col) or None"""
    for merged_range in sheet.merged_cells.ranges:
        if (merged_range.min_row <= row <= merged_range.max_row and
                merged_range.min_col <= col <= merged_range.max_col):
            return (merged_range.min_row, merged_range.min_col,
                    merged_range.max_row, merged_range.max_col)
    return None

# Check if a cell is part of an ignored color range (including merged cells)
# Also helpers to resolve merged top-left value/format

def get_top_left_coords(sheet, row, col):
    merged = get_merged_range(sheet, row, col)
    if merged:
        return merged[0], merged[1]
    return row, col

def get_effective_value(ws_data, ws_format, row, col):
    tl_r, tl_c = get_top_left_coords(ws_format, row, col)
    return ws_data.cell(row=tl_r, column=tl_c).value

def is_cell_in_ignored_range(sheet, row, col):
    """Check if a cell is part of an ignored color range (including merged cells)"""
    if is_cell_in_added_ranges(row, col):
        return False

    cell = sheet.cell(row=row, column=col)
    if is_ignored_color(cell):
        return True

    merged_range = get_merged_range(sheet, row, col)
    if merged_range:
        start_row, start_col, end_row, end_col = merged_range
        top_left_cell = sheet.cell(row=start_row, column=start_col)
        if is_ignored_color(top_left_cell):
            return True

    return False

# Parse a cell range string like 'A1:B5' and return (start_row, start_col, end_row, end_col)
# Uses openpyxl's range_boundaries to handle different formats
# Returns None if the range is invalid or cannot be parsed
def parse_cell_range(cell_range_str):
    """Parse a cell range string like 'A1:B5' and return (start_row, start_col, end_row, end_col)"""
    try:
        from openpyxl.utils import range_boundaries
        min_col, min_row, max_col, max_row = range_boundaries(cell_range_str)
        return (min_row, min_col, max_row, max_col)
    except Exception as e:
        print(f"Error parsing cell range '{cell_range_str}': {e}")
        return None

# Check if a cell is in any of the specified cell ranges to be added
# Uses the parse_cell_range function to handle different formats
def is_cell_in_added_ranges(row, col):
    """Check if a cell is in any of the specified cell ranges to be added"""
    if not CELLS_TO_BE_ADDED or not CELLS_TO_BE_ADDED[0]:
        return False

    for cell_range_str in CELLS_TO_BE_ADDED:
        cell_range_str = cell_range_str.strip()
        if not cell_range_str:
            continue

        parsed_range = parse_cell_range(cell_range_str)
        if parsed_range:
            start_row, start_col, end_row, end_col = parsed_range
            if (start_row <= row <= end_row and start_col <= col <= end_col):
                return True

    return False

# Check if a cell is in any of the specified cell ranges to be ignored
def is_cell_in_ignored_ranges(row, col):
    """Check if a cell is in any of the specified cell ranges to be ignored"""
    if not IGNORED_RANGES or not IGNORED_RANGES[0]:
        return False

    for cell_range_str in IGNORED_RANGES:
        cell_range_str = cell_range_str.strip()
        if not cell_range_str:
            continue

        parsed_range = parse_cell_range(cell_range_str)
        if parsed_range:
            start_row, start_col, end_row, end_col = parsed_range
            if (start_row <= row <= end_row and start_col <= col <= end_col):
                return True

    return False

# Debug function to print cell colors for troubleshooting
def debug_cell_colors(sheet, max_rows=10, max_cols=10):
    """Debug function to print cell colors for troubleshooting"""
    print(f"IGNORED_COLORS: {IGNORED_COLORS}")
    print(f"CELLS_TO_BE_ADDED: {CELLS_TO_BE_ADDED}")
    print(f"IGNORED_RANGES: {IGNORED_RANGES}")
    print("Cell colors in first few rows/columns:")
    for r in range(1, min(max_rows + 1, sheet.max_row + 1)):
        for c in range(1, min(max_cols + 1, sheet.max_column + 1)):
            cell = sheet.cell(row=r, column=c)
            cell_color = normalize_color(cell.fill.start_color)
            if cell_color != "None":
                is_ignored = is_ignored_color(cell)
                is_added = is_cell_in_added_ranges(r, c)
                is_ignored_range = is_cell_in_ignored_ranges(r, c)
                print(
                    f"Cell {cell.coordinate}: color={cell_color}, ignored={is_ignored}, added={is_added}, ignored_range={is_ignored_range}")

# Find all tables in a sheet based on bordered cells
# Returns a list of tuples (start_row, start_col, end_row, end_col)
def find_bordered_tables(sheet):
    visited = set()
    tables = []
    max_row = sheet.max_row
    max_col = sheet.max_column

    def has_border(cell):
        b = cell.border
        return any([b.left.style, b.right.style, b.top.style, b.bottom.style])

    for i in range(1, max_row + 1):
        for j in range(1, max_col + 1):
            if (i, j) in visited:
                continue
            cell = sheet.cell(row=i, column=j)
            if has_border(cell):
                end_row = i
                while end_row + 1 <= max_row and any(has_border(sheet.cell(row=end_row+1, column=k))
                                                     for k in range(j, max_col+1)):
                    end_row += 1
                end_col = j
                while end_col + 1 <= max_col and any(has_border(sheet.cell(row=k, column=end_col+1))
                                                     for k in range(i, end_row+1)):
                    end_col += 1
                for r in range(i, end_row + 1):
                    for c in range(j, end_col + 1):
                        visited.add((r, c))
                tables.append((i, j, end_row, end_col))
    return tables

# Add a hyperlink to the summary sheet
def add_hyperlink(sheet_summary, row, col, target_sheet, display_text=None):
    cell = sheet_summary.cell(row=row, column=col)
    display_text = display_text or target_sheet
    if re.search(r"[^A-Za-z_]", target_sheet):
        safe_sheet_name = f"'{target_sheet}'"
    else:
        safe_sheet_name = target_sheet
    cell.hyperlink = f"#{safe_sheet_name}!A1"
    cell.value = display_text
    cell.font = Font(name="Aptos Narrow", size=11, color="0000FF", underline="single")
    cell.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    cell.alignment = Alignment(horizontal="left", vertical="top")

# Apply header formatting to a cell
def apply_header_formatting(cell):
    cell.font = Font(name="Aptos Narrow", size=12, bold=True, color="000000")
    cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

# Apply data formatting to a cell
def apply_data_formatting(cell):
    cell.font = Font(name="Aptos Narrow", size=11, color="000000")
    cell.alignment = Alignment(horizontal="left", vertical="top")
    cell.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

# --- Main comparison ---
wb_output = Workbook()
summary_ws = wb_output.active
summary_ws.title = "Summary"
summary_ws.append(["Sheet Name", "Hyperlink",
                   "Missing Details", "Issue Count", "Ignored Issues Count"])

# Apply header formatting to summary sheet
for col in range(1, 6):  # 5 columns
    apply_header_formatting(summary_ws.cell(row=1, column=col))
row_summary = 2

# Preserve order: sample sheets first, then IMR-only sheets
ordered_sheetnames = list(wb_sample.sheetnames) + \
    [s for s in wb_imr.sheetnames if s not in wb_sample.sheetnames]

for sheet_name in ordered_sheetnames:
    issues_count = 0
    ignored_issues_count = 0

    if sheet_name not in wb_sample.sheetnames:
        summary_ws.cell(row=row_summary, column=1, value=sheet_name)
        add_hyperlink(summary_ws, row_summary, 2, sheet_name)
        summary_ws.cell(row=row_summary, column=3, value="Missing In Sample")
        summary_ws.cell(row=row_summary, column=4, value=0)
        summary_ws.cell(row=row_summary, column=5, value=0)

        apply_data_formatting(summary_ws.cell(row=row_summary, column=1))
        apply_data_formatting(summary_ws.cell(row=row_summary, column=3))
        apply_data_formatting(summary_ws.cell(row=row_summary, column=4))
        apply_data_formatting(summary_ws.cell(row=row_summary, column=5))
        row_summary += 1
        continue

    if sheet_name not in wb_imr.sheetnames:
        summary_ws.cell(row=row_summary, column=1, value=sheet_name)
        add_hyperlink(summary_ws, row_summary, 2, sheet_name)
        summary_ws.cell(row=row_summary, column=3, value="Missing In IMR_Report")
        summary_ws.cell(row=row_summary, column=4, value=0)
        summary_ws.cell(row=row_summary, column=5, value=0)

        apply_data_formatting(summary_ws.cell(row=row_summary, column=1))
        apply_data_formatting(summary_ws.cell(row=row_summary, column=3))
        apply_data_formatting(summary_ws.cell(row=row_summary, column=4))
        apply_data_formatting(summary_ws.cell(row=row_summary, column=5))
        row_summary += 1
        continue

    ws_sample = wb_sample[sheet_name]
    ws_imr = wb_imr[sheet_name]
    ws_sample_data = wb_sample_data[sheet_name]
    ws_imr_data = wb_imr_data[sheet_name]

    ws_out = None

    def ensure_ws_out(current_ws_out):
        if current_ws_out is None:
            current_ws_out = wb_output.create_sheet(title=sheet_name)
            current_ws_out.append(["Cell", "Name", "Column Name", "Issue Type",
                                   "Sample Value", "Generated Value"])
            for col in range(1, 7):  # 6 columns
                apply_header_formatting(current_ws_out.cell(row=1, column=col))
        return current_ws_out

    # Track ignored issues separately
    ignored_issues = []

    seen_pairs = set()

    # --- Table comparison: include tables from both sheets ---
    table_ranges_sample = find_bordered_tables(ws_sample)
    table_ranges_imr = find_bordered_tables(ws_imr)
    table_cells = set()
    processed_table_cells = set()

    for (start_row, start_col, end_row, end_col) in table_ranges_sample + table_ranges_imr:
        # Mark all cells in this table range as table cells
        for r in range(start_row, end_row + 1):
            for c in range(start_col, end_col + 1):
                table_cells.add((r, c))

        # Process only the data cells (not headers)
        for r in range(start_row + 1, end_row + 1):
            # Prefer sample row header; fallback to IMR (respect merged top-left)
            row_header = get_effective_value(ws_sample_data, ws_sample, r, start_col)
            if row_header is None:
                row_header = get_effective_value(ws_imr_data, ws_imr, r, start_col)

            for c in range(start_col + 1, end_col + 1):
                # Skip cells that are in ignored ranges
                if is_cell_in_ignored_ranges(r, c):
                    continue
                # Avoid duplicate processing across both sheets' table detections
                if (r, c) in processed_table_cells:
                    continue
                processed_table_cells.add((r, c))

                # Prefer sample column header; fallback to IMR (respect merged top-left)
                col_header = get_effective_value(ws_sample_data, ws_sample, start_row, c)
                if col_header is None:
                    col_header = get_effective_value(ws_imr_data, ws_imr, start_row, c)

                # Gate by merged canonical pair to avoid duplicates
                s_r, s_c = get_top_left_coords(ws_sample, r, c)
                i_r, i_c = get_top_left_coords(ws_imr, r, c)
                pair_key = (s_r, s_c, i_r, i_c)
                if pair_key in seen_pairs:
                    continue
                # Only evaluate the top-left of at least one merged region
                if (r != s_r or c != s_c) and (r != i_r or c != i_c):
                    continue
                seen_pairs.add(pair_key)

                # Resolve merged values
                cell_sample_val = get_effective_value(ws_sample_data, ws_sample, r, c)
                cell_imr_val = get_effective_value(ws_imr_data, ws_imr, r, c)

                # Use top-left format cell for merged
                cell_sample_format = ws_sample.cell(row=s_r, column=s_c)
                cell_imr_format = ws_imr.cell(row=i_r, column=i_c)

                is_ignored = is_cell_in_ignored_range(ws_sample, r, c)

                issues = compare_cell(
                    cell_sample_format, cell_imr_format, cell_sample_val, cell_imr_val)
                for issue in issues:
                    if issue["type"] == "Value Mismatch":
                        column_name = col_header
                    else:
                        column_name = safe_str(cell_sample_val)

                    issue_data = [
                        cell_sample_format.coordinate,
                        row_header,
                        column_name,
                        issue["type"],
                        issue["sample"],
                        issue["generated"]
                    ]

                    if is_ignored:
                        ignored_issues.append(issue_data)
                        ignored_issues_count += 1
                    else:
                        ws_out = ensure_ws_out(ws_out)
                        current_row = ws_out.max_row + 1
                        ws_out.append(issue_data)
                        for col in range(1, 7):  # 6 columns
                            apply_data_formatting(ws_out.cell(
                                row=current_row, column=col))
                        issues_count += 1

    # --- Non-table key–value comparison over full union of rows/cols ---
    processed_cells = set()
    max_row = max(ws_sample.max_row, ws_imr.max_row)
    max_col = max(ws_sample.max_column, ws_imr.max_column)

    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):

            if (r, c) in table_cells or (r, c) in processed_cells:
                continue

            if is_cell_in_ignored_ranges(r, c):
                continue

            val_sample = get_effective_value(ws_sample_data, ws_sample, r, c)
            val_imr = get_effective_value(ws_imr_data, ws_imr, r, c)

            if (val_sample and isinstance(val_sample, str) and
                val_sample.strip() and not str(val_sample).strip().isdigit() and
                not val_sample.strip().startswith('(') and
                    len(val_sample.strip()) > 1):

                key = val_sample.strip()

                for cc in range(c + 1, max_col + 1):
                    if (r, cc) in table_cells:
                        break

                    if is_cell_in_ignored_ranges(r, cc):
                        continue

                    value_sample = get_effective_value(ws_sample_data, ws_sample, r, cc)
                    value_imr = get_effective_value(ws_imr_data, ws_imr, r, cc)

                    if value_sample is not None or value_imr is not None:
                        # Gate by merged canonical pair to avoid duplicates
                        s_r, s_c = get_top_left_coords(ws_sample, r, cc)
                        i_r, i_c = get_top_left_coords(ws_imr, r, cc)
                        pair_key = (s_r, s_c, i_r, i_c)
                        if pair_key in seen_pairs:
                            continue
                        if (r != s_r or cc != s_c) and (r != i_r or cc != i_c):
                            continue
                        seen_pairs.add(pair_key)
                        cell_sample_format = ws_sample.cell(row=s_r, column=s_c)
                        cell_imr_format = ws_imr.cell(row=i_r, column=i_c)

                        is_ignored = is_cell_in_ignored_range(ws_sample, r, cc)

                        issues = compare_cell(
                            cell_sample_format, cell_imr_format, value_sample, value_imr)
                        for issue in issues:
                            issue_data = [
                                cell_sample_format.coordinate,
                                key,
                                safe_str(value_sample),
                                issue["type"],
                                issue["sample"],
                                issue["generated"]
                            ]

                            if is_ignored:
                                ignored_issues.append(issue_data)
                                ignored_issues_count += 1
                            else:
                                ws_out = ensure_ws_out(ws_out)
                                current_row = ws_out.max_row + 1
                                ws_out.append(issue_data)
                                for col in range(1, 7):  # 6 columns
                                    apply_data_formatting(ws_out.cell(
                                        row=current_row, column=col))
                                issues_count += 1

                        processed_cells.add((r, cc))
                        break

                processed_cells.add((r, c))

    # --- Fallback sweep: compare any remaining cells outside tables/key-value ---
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            if (r, c) in table_cells or (r, c) in processed_cells:
                continue
            if is_cell_in_ignored_ranges(r, c):
                continue
            # Only evaluate top-left of merged regions (in either sheet) to avoid duplicates
            s_r, s_c = get_top_left_coords(ws_sample, r, c)
            i_r, i_c = get_top_left_coords(ws_imr, r, c)
            pair_key = (s_r, s_c, i_r, i_c)
            if pair_key in seen_pairs:
                continue
            if (r != s_r or c != s_c) and (r != i_r or c != i_c):
                continue
            seen_pairs.add(pair_key)
            val_sample = get_effective_value(ws_sample_data, ws_sample, r, c)
            val_imr = get_effective_value(ws_imr_data, ws_imr, r, c)
            if safe_str(val_sample) == safe_str(val_imr):
                continue
            cell_sample_format = ws_sample.cell(row=s_r, column=s_c)
            cell_imr_format = ws_imr.cell(row=i_r, column=i_c)
            is_ignored = is_cell_in_ignored_range(ws_sample, r, c)
            issues = compare_cell(cell_sample_format, cell_imr_format, val_sample, val_imr)
            for issue in issues:
                issue_data = [
                    ws_sample.cell(row=r, column=c).coordinate,
                    "",
                    "",
                    issue["type"],
                    issue["sample"],
                    issue["generated"]
                ]
                if is_ignored:
                    ignored_issues.append(issue_data)
                    ignored_issues_count += 1
                else:
                    ws_out = ensure_ws_out(ws_out)
                    current_row = ws_out.max_row + 1
                    ws_out.append(issue_data)
                    for col in range(1, 7):  # 6 columns
                        apply_data_formatting(ws_out.cell(row=current_row, column=col))
                    issues_count += 1

    # --- Add Ignored Columns Section ---
    if ignored_issues:
        ws_out = ensure_ws_out(ws_out)
        ws_out.append([])
        ws_out.append(["***Columns To Be Ignored***"])
        apply_header_formatting(ws_out.cell(row=ws_out.max_row, column=1))
        ws_out.append(["Cell", "Name", "Column Name", "Issue Type",
                       "Sample Value", "Generated Value"])
        for col in range(1, 7):  # 6 columns
            apply_header_formatting(ws_out.cell(
                row=ws_out.max_row, column=col))
        for issue_data in ignored_issues:
            current_row = ws_out.max_row + 1
            ws_out.append(issue_data)
            for col in range(1, 7):  # 6 columns
                apply_data_formatting(ws_out.cell(row=current_row, column=col))

    # --- Summary ---
    summary_ws.cell(row=row_summary, column=1, value=sheet_name)
    add_hyperlink(summary_ws, row_summary, 2, sheet_name)
    summary_ws.cell(row=row_summary, column=3, value="")
    summary_ws.cell(row=row_summary, column=4, value=issues_count)
    summary_ws.cell(row=row_summary, column=5, value=ignored_issues_count)

    apply_data_formatting(summary_ws.cell(row=row_summary, column=1))
    apply_data_formatting(summary_ws.cell(row=row_summary, column=3))
    apply_data_formatting(summary_ws.cell(row=row_summary, column=4))
    apply_data_formatting(summary_ws.cell(row=row_summary, column=5))
    row_summary += 1

# Remove default sheet if empty
if "Sheet" in wb_output.sheetnames and wb_output["Sheet"].max_row == 1:
    del wb_output["Sheet"]

# --- Save output ---
wb_output.save(OUTPUT_FILE)
print(f"Comparison report saved to: {OUTPUT_FILE}")